"""Deterministic text extraction from evidence documents.

Every extraction records HOW it was obtained (``method``), because a downstream
release gate must be able to tell a value read straight out of a spreadsheet
cell from one a model guessed at:

    deterministic   parsed from the file's own text/cell layer
    ocr             recovered from pixels (not wired up; declared for honesty)
    model_assisted  produced by an LLM

Images carry no text layer and OCR is not installed, so an image yields a
location with an empty excerpt and ``method="manual_review"``: the operator must
read the certificate and confirm the fact themselves. We never pretend to have
read a JPEG.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger("listing.evidence.extract")

#: How a piece of text was obtained.
METHOD_DETERMINISTIC = "deterministic"
METHOD_OCR = "ocr"
METHOD_MODEL = "model_assisted"
METHOD_MANUAL = "manual_review"

#: Excerpts are capped so the index stays small and no document is mirrored.
MAX_EXCERPT_CHARS = 2000
MAX_LOCATIONS = 200


class Location(dict):
    """One addressable place inside a document, with the text found there."""


def _loc(
    *,
    page: "int | None" = None,
    sheet: str = "",
    cell: str = "",
    excerpt: str = "",
    method: str = METHOD_DETERMINISTIC,
) -> dict[str, Any]:
    return {
        "page": page,
        "sheet": sheet,
        "cell": cell,
        "excerpt": (excerpt or "")[:MAX_EXCERPT_CHARS],
        "method": method,
    }


def _extract_pdf(data: bytes) -> list[dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except Exception:  # pragma: no cover - dependency missing
        logger.warning("pypdf unavailable; pdf stored without a text layer")
        return [_loc(excerpt="", method=METHOD_MANUAL)]
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception:
        # A corrupt or encrypted PDF is not an error the operator can fix here;
        # it just means the text must be confirmed by hand.
        return [_loc(excerpt="", method=METHOD_MANUAL)]

    out: list[dict[str, Any]] = []
    for i, page in enumerate(reader.pages[:MAX_LOCATIONS], start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        if text.strip():
            out.append(_loc(page=i, excerpt=text.strip(), method=METHOD_DETERMINISTIC))
    if not out:
        # A scanned manual has pages but no text layer — say so rather than
        # reporting an empty document.
        out.append(_loc(page=1, excerpt="", method=METHOD_MANUAL))
    return out


def _extract_text(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8", errors="replace").strip()
    return [_loc(excerpt=text, method=METHOD_DETERMINISTIC)] if text else []


def _extract_csv(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        rows = list(csv.reader(io.StringIO(text)))
    except csv.Error:
        return _extract_text(data)
    out: list[dict[str, Any]] = []
    for r, row in enumerate(rows[:MAX_LOCATIONS], start=1):
        line = ", ".join(str(c).strip() for c in row if str(c).strip())
        if line:
            out.append(_loc(sheet="csv", cell=f"row {r}", excerpt=line))
    return out


def _extract_xlsx(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover - dependency missing
        logger.warning("openpyxl unavailable; xlsx stored without a cell layer")
        return [_loc(excerpt="", method=METHOD_MANUAL)]
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return [_loc(excerpt="", method=METHOD_MANUAL)]

    out: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if len(out) >= MAX_LOCATIONS:
                break
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                out.append(
                    _loc(sheet=ws.title, cell=f"row {r}", excerpt=", ".join(cells))
                )
    wb.close()
    return out


def _extract_image(data: bytes) -> list[dict[str, Any]]:
    """Images have no text layer and OCR is not configured.

    We record the pixel dimensions (deterministic, genuinely read from the file)
    but leave the excerpt empty and mark the location manual_review — an image
    can never on its own promote a fact to `verified`.
    """
    dims = ""
    try:
        from PIL import Image

        with Image.open(io.BytesIO(data)) as im:
            dims = f"{im.width}×{im.height}px {im.format or ''}".strip()
    except Exception:
        dims = ""
    return [
        _loc(
            excerpt="",
            method=METHOD_MANUAL,
            **({"cell": dims} if dims else {}),
        )
    ]


_FAMILY_DISPATCH = {
    "pdf": _extract_pdf,
    "text": _extract_text,
    "csv": _extract_csv,
    "xlsx": _extract_xlsx,
    "image": _extract_image,
}


def extract_locations(family: str, data: bytes) -> list[dict[str, Any]]:
    """Addressable text locations for one document. Never raises."""
    fn = _FAMILY_DISPATCH.get(family)
    if fn is None:
        return []
    try:
        return fn(data)[:MAX_LOCATIONS]
    except Exception:  # pragma: no cover - defensive
        logger.warning("extraction failed for family=%s", family)
        return [_loc(excerpt="", method=METHOD_MANUAL)]


def document_text(locations: list[dict[str, Any]]) -> str:
    return "\n".join(str(l.get("excerpt") or "") for l in locations).strip()
