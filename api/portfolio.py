"""Batch portfolio: import many SKUs, then migrate them as one operation.

Phase 1 of the migration engine (``migration.py``) answers "what does this
policy change do to *one* SKU's artifacts?". This module lifts the same
deterministic machinery to a portfolio: import a spreadsheet of SKUs, compute
the blast radius across all of them, shadow-compile candidate patches, approve
the safe ones in bulk, and roll back either a single SKU or the whole batch.

Nothing here calls a model — it reuses ``migration``'s deterministic path — and
nothing here publishes anywhere.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import datetime, timezone
from typing import Any

import migration
import skufacts
from checker import apply_checks
from drafts import fallback_drafts

PLATFORMS = ("amazon", "tiktok", "shopify")

#: Row statuses in the portfolio matrix.
UNAFFECTED = "unaffected"
SAFE_PATCH = "safe_patch"
REVIEW_REQUIRED = "review_required"
BLOCKED = "blocked"
APPLIED = "applied"
ROLLED_BACK = "rolled_back"

TEMPLATE_COLUMNS = ["sku", "product_name", "selling_points", "platforms", "evidence_sources"]

TEMPLATE_CSV = (
    "sku,product_name,selling_points,platforms,evidence_sources\n"
    "AERO-350,AeroFold Collapsible Silicone Travel Cup,"
    '"折叠到 4cm|食品级硅胶|防漏盖，350ml",amazon;tiktok;shopify,\n'
    "AERO-500,AeroFold Collapsible Silicone Travel Cup 500ml,"
    '"折叠到 5cm|食品级硅胶|防漏盖，500ml",amazon;shopify,\n'
)


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# --------------------------------------------------------------------------- #
# Import                                                                       #
# --------------------------------------------------------------------------- #


def _split_points(raw: str) -> str:
    """Selling points arrive pipe- or newline-separated; store one per line."""
    parts = [p.strip() for p in re.split(r"[|\n]", raw or "") if p.strip()]
    return "\n".join(parts)


def _parse_platforms(raw: str) -> "tuple[list[str], str]":
    if not (raw or "").strip():
        return list(PLATFORMS), ""
    wanted = [p.strip().lower() for p in re.split(r"[;,]", raw) if p.strip()]
    good = [p for p in wanted if p in PLATFORMS]
    bad = [p for p in wanted if p not in PLATFORMS]
    if not good:
        return [], f"平台无法识别：{'、'.join(bad)}"
    note = f"忽略无法识别的平台：{'、'.join(bad)}" if bad else ""
    return good, note


def _rows_from_csv(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [
        {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        for row in csv.DictReader(io.StringIO(text))
    ]


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c or "").strip().lower() for c in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        out.append(
            {
                header[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                for i in range(len(header))
            }
        )
    return out


def parse_portfolio(data: bytes, family: str) -> dict[str, Any]:
    """Parse an uploaded portfolio file into SKU rows plus a validation report.

    A malformed row is reported and skipped; it never aborts the import, so one
    bad line in a 500-row sheet cannot cost the operator the other 499.
    """
    try:
        raw_rows = _rows_from_xlsx(data) if family == "xlsx" else _rows_from_csv(data)
    except Exception as exc:  # pragma: no cover - corrupt upload
        return {
            "skus": [],
            "errors": [{"row": 0, "sku": "", "error": f"文件无法解析（{type(exc).__name__}）。"}],
            "summary": {"total_rows": 0, "imported": 0, "rejected": 0},
        }

    skus: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()

    for i, row in enumerate(raw_rows, start=2):  # row 1 is the header
        sku = (row.get("sku") or "").strip()
        name = (row.get("product_name") or row.get("name") or "").strip()
        points_raw = row.get("selling_points") or row.get("points") or ""

        if not any([sku, name, points_raw.strip()]):
            continue  # a genuinely blank line is not an error

        if not sku:
            errors.append({"row": i, "sku": "", "error": "缺少 sku 列。"})
            continue
        if sku in seen:
            errors.append({"row": i, "sku": sku, "error": "SKU 重复，已跳过。"})
            continue
        if not name:
            errors.append({"row": i, "sku": sku, "error": "缺少 product_name。"})
            continue

        points = _split_points(points_raw)
        if not points:
            errors.append({"row": i, "sku": sku, "error": "缺少 selling_points。"})
            continue

        platforms, note = _parse_platforms(row.get("platforms") or "")
        if not platforms:
            errors.append({"row": i, "sku": sku, "error": note or "没有可用平台。"})
            continue
        if note:
            errors.append({"row": i, "sku": sku, "error": note, "severity": "warning"})

        evidence_refs = [
            r.strip()
            for r in re.split(r"[;,]", row.get("evidence_sources") or "")
            if r.strip()
        ]

        skus.append(
            {
                "sku": sku,
                "product_name": name,
                "points": points,
                "platforms": platforms,
                "evidence_sources": evidence_refs,
                "row": i,
            }
        )
        seen.add(sku)

    return {
        "skus": skus,
        "errors": errors,
        "summary": {
            "total_rows": len(raw_rows),
            "imported": len(skus),
            "rejected": len([e for e in errors if e.get("severity") != "warning"]),
        },
    }


# --------------------------------------------------------------------------- #
# Artifacts + blast radius                                                     #
# --------------------------------------------------------------------------- #


def _sku_title(sku: dict[str, Any], platform: str) -> str:
    """A per-SKU title derived deterministically from the row's own data.

    ``fallback_drafts`` is a fixed demo product, so every portfolio row would
    otherwise compile to the identical title and a policy change would be
    all-or-nothing across the portfolio. Building the title from the row's
    product name and its own selling points makes the portfolio genuinely
    heterogeneous, which is what a blast radius is supposed to discriminate.
    """
    name = sku["product_name"]
    points = [p for p in sku["points"].split("\n") if p.strip()]
    # the shortest couple of points read as attributes rather than sentences
    attrs = [p.strip() for p in sorted(points, key=len)[:2]]
    if platform == "shopify":
        return name
    return ", ".join([name, *attrs]) if attrs else name


def compile_sku(sku: dict[str, Any]) -> list[dict[str, Any]]:
    """Deterministic artifacts for one imported SKU.

    Uses the same ``apply_checks`` grading path as single-SKU generation, so a
    portfolio row is judged by exactly the rules a hand-entered SKU is. The
    body copy comes from the shared fallback drafts; the title is per-SKU.
    No model is called.
    """
    drafts = fallback_drafts(
        sku["product_name"], sku["points"], "compliant", list(sku["platforms"])
    )
    for draft in drafts:
        draft["title"] = _sku_title(sku, draft["id"])
    out: list[dict[str, Any]] = []
    for draft in drafts:
        graded = apply_checks(
            draft,
            product_name=sku["product_name"],
            points=sku["points"],
            asset_mode="compliant",
        )
        out.append(
            {
                "artifact_id": f"{sku['sku']}::{graded['id']}",
                "sku": sku["sku"],
                "platform": graded["id"],
                "kind": "listing",
                "revision": 1,
                "status": "current",
                "policy_version": graded.get("policyVersion", ""),
                "sku_revision": graded.get("skuRevision", ""),
                "title": graded["title"],
                "title_fact_refs": graded.get("titleFactRefs", []),
                "fields": [
                    {
                        "name": f.get("field", ""),
                        "label": f.get("label", ""),
                        "value": f.get("value", ""),
                        "fact_refs": f.get("factRefs", []),
                    }
                    for f in graded.get("fields", [])
                ],
            }
        )
    return out


def _severity_for(row: dict[str, Any], patch: "dict[str, Any] | None") -> str:
    if patch is None:
        return REVIEW_REQUIRED if row.get("fields_to_regenerate") else UNAFFECTED
    if patch.get("needs_human_review") or not patch.get("validation", {}).get("ok", True):
        return REVIEW_REQUIRED
    return SAFE_PATCH


def analyze_portfolio(
    skus: list[dict[str, Any]],
    *,
    base_policy_version: "str | None" = None,
    candidate_policy_version: "str | None" = None,
    points_override: "dict[str, str] | None" = None,
) -> dict[str, Any]:
    """Blast radius across the whole portfolio, as a SKU × platform × field matrix.

    ``points_override`` maps sku -> new selling points, for a product-fact drift
    that touches only part of the portfolio.
    """
    points_override = points_override or {}
    matrix: list[dict[str, Any]] = []
    per_sku: list[dict[str, Any]] = []
    all_artifacts: dict[str, list[dict[str, Any]]] = {}

    affected_platforms: set[str] = set()
    affected_fields: set[str] = set()

    for sku in skus:
        artifacts = compile_sku(sku)
        all_artifacts[sku["sku"]] = artifacts

        facts_before = skufacts.parse_sku_facts(sku["product_name"], sku["points"])
        new_points = points_override.get(sku["sku"], sku["points"])
        facts_after = skufacts.parse_sku_facts(sku["product_name"], new_points)

        impact = migration.analyze_impact(
            artifacts,
            facts_before=facts_before,
            facts_after=facts_after,
            base_policy_version=base_policy_version,
            candidate_policy_version=candidate_policy_version,
        )

        candidate = migration.build_candidate_patches(
            artifacts,
            impact,
            facts_before=facts_before,
            facts_after=facts_after,
            base_policy_version=base_policy_version,
            candidate_policy_version=candidate_policy_version,
        )
        patch_by_target = {(p["artifact_id"], p["field"]): p for p in candidate["patches"]}

        sku_rows: list[dict[str, Any]] = []
        for row in impact["affected"]:
            platform = row["platform"]
            affected_platforms.add(platform)
            for field in row["fields_to_regenerate"]:
                affected_fields.add(field)
                patch = patch_by_target.get((row["artifact_id"], field))
                status = _severity_for(row, patch)
                reason = "；".join(r["detail"] for r in row["reasons"]) or "需重编译"
                sku_rows.append(
                    {
                        "sku": sku["sku"],
                        "platform": platform,
                        "field": field,
                        "artifact_id": row["artifact_id"],
                        "status": status,
                        "cause": row["cause"],
                        "reason": reason,
                        "previous_value": (patch or {}).get("previous_value", ""),
                        "candidate_value": (patch or {}).get("candidate_value", ""),
                        "note": (patch or {}).get("note", ""),
                    }
                )
            if not row["fields_to_regenerate"]:
                # affected but nothing to recompile — re-validation only
                sku_rows.append(
                    {
                        "sku": sku["sku"], "platform": platform, "field": "-",
                        "artifact_id": row["artifact_id"], "status": UNAFFECTED,
                        "cause": row["cause"],
                        "reason": "；".join(r["detail"] for r in row["reasons"]),
                        "previous_value": "", "candidate_value": "", "note": "",
                    }
                )

        for row in impact["unaffected"]:
            sku_rows.append(
                {
                    "sku": sku["sku"], "platform": row["platform"], "field": "-",
                    "artifact_id": row["artifact_id"], "status": UNAFFECTED,
                    "cause": None, "reason": row.get("reason", "无依赖重叠。"),
                    "previous_value": "", "candidate_value": "", "note": "",
                }
            )

        matrix.extend(sku_rows)
        touched = [r for r in sku_rows if r["status"] != UNAFFECTED]
        per_sku.append(
            {
                "sku": sku["sku"],
                "product_name": sku["product_name"],
                "affected": bool(touched),
                "safe_patch": sum(1 for r in touched if r["status"] == SAFE_PATCH),
                "review_required": sum(1 for r in touched if r["status"] == REVIEW_REQUIRED),
                "blocked": sum(1 for r in touched if r["status"] == BLOCKED),
                "platforms": sorted({r["platform"] for r in touched}),
            }
        )

    affected_skus = [s for s in per_sku if s["affected"]]
    return {
        "generated_at": _now(),
        "matrix": matrix,
        "per_sku": per_sku,
        "artifacts": all_artifacts,
        "summary": {
            "skus_scanned": len(skus),
            "skus_affected": len(affected_skus),
            "skus_unaffected": len(per_sku) - len(affected_skus),
            "affected_platforms": sorted(affected_platforms),
            "affected_fields": sorted(affected_fields),
            "safe_patch": sum(1 for r in matrix if r["status"] == SAFE_PATCH),
            "review_required": sum(1 for r in matrix if r["status"] == REVIEW_REQUIRED),
            "blocked": sum(1 for r in matrix if r["status"] == BLOCKED),
            "unaffected_rows": sum(1 for r in matrix if r["status"] == UNAFFECTED),
        },
        "policy": {
            "base_version": base_policy_version,
            "candidate_version": candidate_policy_version,
        },
    }


# --------------------------------------------------------------------------- #
# Batch apply + rollback                                                       #
# --------------------------------------------------------------------------- #


def snapshot_portfolio(artifacts_by_sku: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    payload = {sku: [dict(a) for a in arts] for sku, arts in artifacts_by_sku.items()}
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return {
        "snapshot_id": hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16],
        "taken_at": _now(),
        "artifacts": payload,
    }


def apply_batch(
    artifacts_by_sku: dict[str, list[dict[str, Any]]],
    approved: list[dict[str, Any]],
    *,
    candidate_policy_version: "str | None" = None,
) -> dict[str, Any]:
    """Apply approved patches per SKU.

    A risky patch cannot slip through here: rows marked ``review_required`` are
    rejected outright rather than quietly applied, so bulk approval of the safe
    set can never widen into approving the rest.
    """
    by_sku: dict[str, list[dict[str, Any]]] = {}
    rejected: list[dict[str, Any]] = []
    for patch in approved:
        if patch.get("status") == REVIEW_REQUIRED or patch.get("needs_human_review"):
            rejected.append(
                {
                    "sku": patch.get("sku"),
                    "artifact_id": patch.get("artifact_id"),
                    "field": patch.get("field"),
                    "reason": "该补丁标记为需人工复核，不能通过批量批准应用。",
                }
            )
            continue
        by_sku.setdefault(str(patch.get("sku")), []).append(patch)

    results: dict[str, Any] = {}
    applied_skus: list[str] = []
    review_skus: list[str] = []

    for sku, artifacts in artifacts_by_sku.items():
        patches = by_sku.get(sku, [])
        if not patches:
            results[sku] = {"artifacts": [dict(a) for a in artifacts], "applied": [], "review": []}
            continue
        out = migration.apply_patches(
            artifacts, patches, candidate_policy_version=candidate_policy_version
        )
        results[sku] = {
            "artifacts": out["artifacts"],
            "applied": out["applied_artifact_ids"],
            "review": out["needs_human_review_ids"],
        }
        if out["applied_artifact_ids"]:
            applied_skus.append(sku)
        if out["needs_human_review_ids"]:
            review_skus.append(sku)

    return {
        "generated_at": _now(),
        "results": results,
        "applied_skus": sorted(applied_skus),
        "needs_review_skus": sorted(review_skus),
        "rejected_patches": rejected,
    }


def rollback_batch(
    snapshot: dict[str, Any], *, only_sku: "str | None" = None
) -> dict[str, Any]:
    """Restore the whole batch, or one SKU from it. Deterministic, no model."""
    if not isinstance(snapshot, dict) or not isinstance(snapshot.get("artifacts"), dict):
        raise ValueError("rollback 需要包含 'artifacts' 的批次快照。")
    stored = snapshot["artifacts"]
    if only_sku is not None:
        if only_sku not in stored:
            raise ValueError(f"快照中没有 SKU {only_sku}。")
        restored = {only_sku: [dict(a) for a in stored[only_sku]]}
    else:
        restored = {sku: [dict(a) for a in arts] for sku, arts in stored.items()}
    return {
        "generated_at": _now(),
        "restored_from": snapshot.get("snapshot_id", ""),
        "scope": only_sku or "batch",
        "artifacts": restored,
    }


# --------------------------------------------------------------------------- #
# Audit report                                                                 #
# --------------------------------------------------------------------------- #


def build_batch_report(
    *,
    analysis: dict[str, Any],
    apply_result: "dict[str, Any] | None" = None,
    status: str = "candidate",
    approver: str = "",
    evidence_versions: "list[dict[str, Any]] | None" = None,
    rollback: "dict[str, Any] | None" = None,
) -> dict[str, Any]:
    matrix = analysis.get("matrix", [])
    applied_skus = set((apply_result or {}).get("applied_skus", []))

    patched = [
        {
            "sku": r["sku"],
            "platform": r["platform"],
            "field": r["field"],
            "original_value": r["previous_value"],
            "patched_value": r["candidate_value"],
            "status": r["status"],
            "reason": r["reason"],
            "applied": r["sku"] in applied_skus and r["status"] == SAFE_PATCH,
        }
        for r in matrix
        if r["status"] in (SAFE_PATCH, REVIEW_REQUIRED)
    ]

    return {
        "schema": "listing-batch-migration-report/v1",
        "generated_at": _now(),
        "status": status,
        "approver": approver,
        "policy": analysis.get("policy", {}),
        "evidence_versions": evidence_versions or [],
        "summary": analysis.get("summary", {}),
        "per_sku": analysis.get("per_sku", []),
        "patched_fields": patched,
        "preserved_rows": [
            {"sku": r["sku"], "platform": r["platform"]}
            for r in matrix
            if r["status"] == UNAFFECTED
        ],
        "validation": {
            "review_required": [r for r in matrix if r["status"] == REVIEW_REQUIRED],
            "blocked": [r for r in matrix if r["status"] == BLOCKED],
        },
        "apply": {
            "applied_skus": sorted(applied_skus),
            "needs_review_skus": (apply_result or {}).get("needs_review_skus", []),
            "rejected_patches": (apply_result or {}).get("rejected_patches", []),
        },
        "rollback": rollback or {},
    }


def render_batch_report_html(report: dict[str, Any]) -> str:
    esc = lambda s: (  # noqa: E731
        str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )
    su = report.get("summary", {})
    rows = "".join(
        f"<tr><td>{esc(p['sku'])}</td><td>{esc(p['platform'])}</td><td>{esc(p['field'])}</td>"
        f"<td>{esc(p['original_value'])}</td><td>{esc(p['patched_value'])}</td>"
        f"<td>{esc(p['status'])}</td><td>{'✓' if p['applied'] else '—'}</td></tr>"
        for p in report.get("patched_fields", [])
    )
    pol = report.get("policy", {})
    return f"""<!doctype html><html lang="zh-CN"><meta charset="utf-8">
<title>批量迁移审计报告</title>
<style>body{{font:14px/1.6 system-ui,sans-serif;max-width:1100px;margin:2rem auto;padding:0 1rem}}
table{{border-collapse:collapse;width:100%;margin:1rem 0}}
td,th{{border:1px solid #ddd;padding:.4rem .6rem;text-align:left;vertical-align:top}}
.meta{{color:#666}}</style>
<h1>批量迁移审计报告 · {esc(report.get('status',''))}</h1>
<p class="meta">{esc(report.get('generated_at',''))}
{' · 批准人 ' + esc(report['approver']) if report.get('approver') else ''}</p>
<p>政策：{esc(pol.get('base_version','—'))} → {esc(pol.get('candidate_version','—'))}</p>
<p>扫描 SKU {su.get('skus_scanned',0)} · 受影响 {su.get('skus_affected',0)} ·
未受影响 {su.get('skus_unaffected',0)} · 可安全修补 {su.get('safe_patch',0)} ·
需人工 {su.get('review_required',0)} · 阻断 {su.get('blocked',0)}</p>
<table><thead><tr><th>SKU</th><th>平台</th><th>字段</th><th>原值</th><th>补丁值</th>
<th>状态</th><th>已应用</th></tr></thead>
<tbody>{rows or '<tr><td colspan="7">无</td></tr>'}</tbody></table>
<p class="meta">本报告不代表任何平台发布动作，本工具不做自动上架。</p>
</html>"""
